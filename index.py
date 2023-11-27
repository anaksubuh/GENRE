import base64
import json
import time
import telebot
import datetime
import openpyxl
import pandas as pd
from PIL import Image
import streamlit as st
import streamlit.components.v1 as components
from streamlit_option_menu import option_menu
from streamlit_extras.switch_page_button import switch_page

import openpyxl
wb = openpyxl.Workbook()
ws = wb.active

# modul telegram untuk mengecek id
api_telegram_bot = '6215631300:AAH0rPZsMxTT_fnEva2sjGeVqugTdQwk_t8'
bot = telebot.TeleBot(api_telegram_bot)
@bot.message_handler(commands=['ping'])
def greet(message):
    print(message.chat.id)
    bot.reply_to(message, f"Hey! , Thankyou for try your verification\nYour code is :\n{message.chat.id}")

chat_id = -4029218509

kirim_tele = True

st.set_page_config(
    page_title='GENRE KOTA MAGELANG',
    page_icon='logo.png',
    layout='wide',  
    initial_sidebar_state='expanded',
    menu_items={
        'Get Help': 'https://www.extremelycoolapp.com/help',
        'Report a bug': 'https://www.extremelycoolapp.com/bug',
         'About': '# This is a header. This is an *extremely* cool app!'
    }
)

with st.sidebar:

    st.image('logo.png',width=95)
    selected = option_menu('Menu', ['Dasboard','Curhat Kuy','Curhatku','Chat live','Setting'], icons=['house','book','list','chat','gear'], menu_icon="cast", default_index=0)

if selected == 'Dasboard':

    hide_streamlit_style = """
                <style>
                    #MainMenu {visibility: hidden;}
                    #stToolbar {visibility: hidden;}
                    #viewerBadge_container__r5tak styles_viewerBadge__CvC9N {visibility: hidden;}
                    footer {visibility: hidden;}
                </style>
                """
    st.markdown(hide_streamlit_style, unsafe_allow_html=True) 

    st.markdown(
        """
        <style>
        .css-1jc7ptx, .e1ewe7hr3, .viewerBadge_container__1QSob,
        .styles_viewerBadge__1yB5_, .viewerBadge_link__1S137,
        .viewerBadge_text__1JaDK {
            display: none;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    col1, col2, col3 = st.columns(3)
    with col2:
        st.image('logo.png',width=300)
    
    st.header('Curhat via website :')
    st.write('-Hallo Genrengers')
    st.write('-Kamu bisa curhat sebanyak seaman dan senyaman mungkin di sini lho gaess, gimana caranya?.')
    st.write('-Langsung aja klik (>) di atas kiri lalu pilih bagian (Curhat kuy).')
    st.write('-Ketika kamu sudah masuk ke halaman (Curhat kuy) kamu langsung bisa isi form curhatmu.')
    st.write('-Seberapa aman sih NGL genre ini? , Jika kamu tidak nyaman memberi tahu nama kalian kalian cukup mengosongkan bagian nama itu karena optional (boleh iya maupun tidak).')
    st.write('-cara privacy nama dan cerita')
    st.write('- Jika namamu tidak ingin di ketahui siapapun , mohon tidak mengisi kolom nama')
    st.write('- Jika cerita dan namamu hanya untuk genre , mohon checklist (Hanya admin GENRE yang bisa melihat.).')
    st.header('')
    
    st.header('Konseling via instagram & wa :')
    st.write('- Kamu bisa curhat sebanyak seaman dan senyaman mungkin di sini lho gaess, gimana caranya?')
    st.write('- Langsung aja klik (>) di atas kiri lalu pilih bagian (Chat Live)')
    st.write('- Okei di situlah kamu bisa memilih ingin kontak genre atau konseling langsung dari whatsapp maupun instagram.')

    for i in range(5):
        st.write('')

    col1, col2, col3 = st.columns(3)
    with col1:
        pass
    with col2:
        st.write('© COPYRIGHT 2023 - GENERASI BERENCANA.')
        st.write('ALL RIGHTS RESERVED. WEBSITE DESIGN BY.')
        st.link_button('GENRE MGL','https://www.instagram.com/genrekotamagelang/')
        pass
    with col3:
        pass

if selected == 'Curhat Kuy':

    hide_streamlit_style = """
                <style>
                    #MainMenu {visibility: hidden;}
                    #stToolbar {visibility: hidden;}
                    #viewerBadge_container__r5tak styles_viewerBadge__CvC9N {visibility: hidden;}
                    footer {visibility: hidden;}
                </style>
                """
    st.markdown(hide_streamlit_style, unsafe_allow_html=True) 

    col1, col2, col3 = st.columns(3)
    with col2:
        st.image('logo.png',width=300)

    nama   = st.text_input('Name (optional) :','')
    age = st.slider('How old are you?', 0, 100, 17)
    st.write("I'm ", age, 'years old')  
    curhat = st.text_area('TULISKAN CERITA HATIMU HARI INI...')

    publick = st.checkbox('Hanya admin GENRE yang bisa melihat.')

    if st.button('Submit'):

        if nama or curhat != str():

            a = open(f'history.json')
            data = json.load(a)
            x = curhat
            y = 'skip...'
            if x in data:
                #print(f'[+] chek : [{x}] | {data[x]}')
                st.warning('[+] Pesan mu sudah terkirim , silahkan chek pada CURHATKU')
            else:
                data[x] = y
                #print(f'[+] chek : [{x}] | save..')
                b = open(f'history.json', 'w')
                json.dump(data,b)
                b.close()

                current_time = datetime.datetime.now()
                now = current_time.strftime('%H:%M:%S')
                date = current_time.strftime('%d/%m/%Y')

                if publick:

                    st.warning('[+] Pesanmu sudah terkirim , Hanya admin GENRE yang bisa melihat.')
                    # write database normal
                    f = open('cerita_global.txt','a')
                    f.write(f'Name : {nama}|age : {age}|post date : {now} {date}|Hidden : True\n{curhat}\n')
                    f.close()

                    # sending master database
                    f = open('cerita_global.txt','r', encoding="utf-8")
                    data_export_tele = f.read()
                    for i in range(kirim_tele):
                        bot = telebot.TeleBot(api_telegram_bot)
                        bot.send_message(chat_id,f'NGL GENRE...\n{data_export_tele}')

                else:

                    st.warning('[+] Chek pesanmu pada page Curhatku pada taskbar')
                    # write database normal
                    f = open('cerita_publick.txt','a')
                    f.write(f'Name : {nama}|age : {age} {date}\n{curhat}\n')
                    f.close()

                    # write database normal
                    f = open('cerita_global.txt','a')
                    f.write(f'Name : {nama}|age : {age}|post date : {now} {date}|Hidden : False\n{curhat}\n')
                    f.close()

                    # sending master database
                    f = open('cerita_global.txt','r', encoding="utf-8")
                    data_export_tele = f.read()
                    for i in range(kirim_tele):
                        bot = telebot.TeleBot(api_telegram_bot)
                        bot.send_message(chat_id,f'NGL GENRE...\n{data_export_tele}')

            
        else:
            st.warning('[+] kamu tidak membuat pesan')
            
            time.sleep(2)

    for i in range(5):
        st.write('')

    col1, col2, col3 = st.columns(3)
    with col1:
        pass
    with col2:
        st.write('© COPYRIGHT 2023 - GENERASI BERENCANA.')
        st.write('ALL RIGHTS RESERVED. WEBSITE DESIGN BY.')
        st.link_button('GENRE MGL','https://www.instagram.com/genrekotamagelang/')
        pass
    with col3:
        pass

if selected == 'Curhatku':

    hide_streamlit_style = """
                <style>
                    #MainMenu {visibility: hidden;}
                    #stToolbar {visibility: hidden;}
                    #viewerBadge_container__r5tak styles_viewerBadge__CvC9N {visibility: hidden;}
                    footer {visibility: hidden;}
                </style>
                """
    st.markdown(hide_streamlit_style, unsafe_allow_html=True) 

    col1, col2, col3 = st.columns(3)
    with col1:
        pass
    with col2:
        st.image('logo.png',width=300)
        pass
    with col3:
        pass

    # reading database
    database = open('cerita_publick.txt', 'r', encoding="utf-8")
    count = 0
    # Using for loop
    for data in database:
        count += 1
        text = (str(data.strip()))
        ganjil_genap = ('genap' if (count % 2 == 0) else 'ganjil')
        if ganjil_genap == 'ganjil':
            col1,col2= st.columns(2)
            with col1:
                st.success(text, icon="ℹ")
            with col2:
                pass

        if ganjil_genap == 'genap':
            st.info(text, icon="ℹ")
            st.warning('|'+'-'*60+'|')
        st.caption('')

    for i in range(5):
        st.write('')

    col1, col2, col3 = st.columns(3)
    with col1:
        pass
    with col2:
        st.write('© COPYRIGHT 2023 - GENERASI BERENCANA.')
        st.write('ALL RIGHTS RESERVED. WEBSITE DESIGN BY.')
        st.link_button('GENRE MGL','https://www.instagram.com/genrekotamagelang/')
        pass
    with col3:
        pass

if selected == 'Chat live':

    hide_streamlit_style = """
                <style>
                    #MainMenu {visibility: hidden;}
                    #stToolbar {visibility: hidden;}
                    #viewerBadge_container__r5tak styles_viewerBadge__CvC9N {visibility: hidden;}
                    footer {visibility: hidden;}
                </style>
                """
    st.markdown(hide_streamlit_style, unsafe_allow_html=True) 

    st.header('Hallo sobat GENRE INDOENSIA')
    st.write('Temen-temen pernah gak sih mau cerita tapi gak tau harus cerita ke mana, karena tidak ada tempat bercerita yang aman dan nyaman :( ')
    st.write('Kami genre kota magelang menyediakan tempat pelayanan dengan basis konseling sebaya atau curhat secara aman nyaman baik public maupun private😃.')
    st.write('Loh kak bagaimana caranya? , teman-teman langsung bisa chat secara private dengan cara click kolom di bawah ini 😊.')

    col1, col2, col3, col4, col5, col6, col7, col8, col9= st.columns(9)
    with col1:
        st.write('')
        st.link_button('whatsapp 💬','https://wa.me/6282331970107?text=hallo%20genre%20kota%20magelang%20:)')
    with col2:
        st.write('')
        st.link_button('Instagram 💬','https://www.instagram.com/genrekotamagelang/')
    #with col3:
        #st.link_button('Twitter 💬','https://twitter.com/genre_indonesia')

    # chat auto balas
    #prompt = st.chat_input("Say something")
    #if prompt:
    #    st.write(f"ROMEO : {prompt}")

    for i in range(5):
        st.write('')

    col1, col2, col3 = st.columns(3)
    with col1:
        pass
    with col2:
        st.write('© COPYRIGHT 2023 - GENERASI BERENCANA.')
        st.write('ALL RIGHTS RESERVED. WEBSITE DESIGN BY.')
        st.link_button('GENRE MGL','https://www.instagram.com/genrekotamagelang/')
        pass
    with col3:
        pass

if selected == 'Setting':

    hide_streamlit_style = """
                <style>
                    #MainMenu {visibility: hidden;}
                    #stToolbar {visibility: hidden;}
                    #viewerBadge_container__r5tak styles_viewerBadge__CvC9N {visibility: hidden;}
                    footer {visibility: hidden;}
                </style>
                """
    st.markdown(hide_streamlit_style, unsafe_allow_html=True) 

    # Security
    #passlib,hashlib,bcrypt,scrypt
    import hashlib
    def make_hashes(password):
        return hashlib.sha256(str.encode(password)).hexdigest()

    def check_hashes(password,hashed_text):
        if make_hashes(password) == hashed_text:
            return hashed_text
        return False
    # DB Management
    import sqlite3 
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    # DB  Functions
    def create_usertable():
        c.execute('CREATE TABLE IF NOT EXISTS userstable(username TEXT,password TEXT)')

    def add_userdata(username,password):
        c.execute('INSERT INTO userstable(username,password) VALUES (?,?)',(username,password))
        conn.commit()

    def login_user(username,password):
        c.execute('SELECT * FROM userstable WHERE username =? AND password = ?',(username,password))
        data = c.fetchall()
        return data

    def view_all_users():
        c.execute('SELECT * FROM userstable')
        data = c.fetchall()
        return data

    def main():
        """Simple Login App"""

        menu = ["Login","SignUp"]
        choice = st.sidebar.selectbox("Menu",menu)

        if choice == "Login":
            st.subheader("Login/Logout")

            username = st.text_input("User Name")
            password = st.text_input("Password",type='password')
            if st.checkbox("Login"):
                # if password == '12345':
                create_usertable()
                hashed_pswd = make_hashes(password)

                result = login_user(username,check_hashes(password,hashed_pswd))
                if result:

                    st.success("Logged In as {}".format(username))

                    task = st.selectbox("Task",["SUARA HATI","Analytics","Profiles"])
                    if task == "SUARA HATI":

                        # ============== MAKE DATA TXT TO EXEL ==============#
                        # Using readlines()
                        file1 = open('cerita_global.txt', 'r', encoding="utf-8")
                        Lines = file1.readlines()
                        count = 0
                        # Strips the newline character
                        for line in Lines:
                            count += 1
                            linee = count
                            data = (count, line.strip())
                            data = str(data)
                            tasta = ('IDENTITAS' if (count % 2 == 0) else 'CERITA')

                            if tasta == 'CERITA':

                                # data remove
                                data = data.replace(f"({count}, 'Name : "," ")
                                data = data.replace("age : "," ")
                                data = data.replace("post date : "," ")
                                data = data.replace("Hidden : "," ")
                                data = data.replace("')"," ")
                                data = data.split('|')
                                #============== ==============#
                                nama     = data[0]
                                age      = data[1]
                                postdate = data[2]
                                hidden   = data[3]

                                ws[f'A{count}'] = nama
                                ws[f'B{count}'] = age
                                ws[f'C{count}'] = postdate
                                ws[f'D{count}'] = hidden

                            if tasta == "IDENTITAS":
                                data = data.replace(f"({count}, '"," ")
                                data = data.replace("')"," ")
                                data = str(data)
                                cerita = data

                                #print(f'Name      : {nama}\nig        : {ig}\nage       : {age}\npost date : {postdate}\nHidden    : {hidden}\ncerita    : {cerita}')
                                #print('')
                                ws[f'F{count}'] = cerita

                        wb.save('genre.csv')
                        # ============== MAKE DATA TXT TO EXEL ==============#

                        col1, col2, col3, col4, col5, col6= st.columns(6)
                        with col1:
                            st.info('Import All data user in csv : ', icon="ℹ️")
                        with col2:
                            with open("genre.csv", "rb") as file:
                                btn = st.download_button(
                                        label="DOWNLOAD genre.csv",
                                        data=file,
                                        file_name="genre.csv",
                                        mime=""
                                    )

                        # reading database
                        database = open('cerita_global.txt', 'r', encoding="utf-8")
                        count = 0
                        # Using for loop
                        for data in database:
                            count += 1
                            text = (str(data.strip()))
                            ganjil_genap = ('genap' if (count % 2 == 0) else 'ganjil')
                            if ganjil_genap == 'ganjil':
                                col1, col2= st.columns(2)
                                with col1:
                                    st.success(text, icon="ℹ")
                                    if text.count('Hidden : True'):
                                        text = text.replace('200'or'404','')
                                        status_hidden = True
                                    if text.count('Hidden : False'):
                                        status_hidden = False
                                with col2:
                                    st.info(f"Hidden : {status_hidden}")

                            if ganjil_genap == 'genap':
                                if status_hidden == True:
                                    st.error(text, icon="🚨")
                                else:
                                    st.info(text, icon="ℹ️")
                                st.warning('|'+'-'*60+'|')
                            st.caption('')

                    elif task == "Analytics":
                        st.caption('BELUM ADA KONTEN')
                    elif task == "Profiles":
                        st.caption('BELUM ADA KONTEN')
                        st.subheader("User Profiles")
                        user_result = view_all_users()
                        clean_db = pd.DataFrame(user_result,columns=["Username","Password"])
                        st.dataframe(clean_db)
                else:
                    st.warning("Incorrect Username/Password")

        elif choice == "SignUp":
            st.subheader("Create New Account")
            new_user = st.text_input("Username")
            new_password = st.text_input("Password",type='password')
            code = st.text_input('CODE VERIFIKATOR')

            if st.button("Signup"):
                if code == '2801':
                    create_usertable()
                    add_userdata(new_user,make_hashes(new_password))
                    st.success("You have successfully created a valid Account")
                    st.info("Go to Login Menu to login")
                else:
                    st.warning('[+] CODE salah , mohon hubungi developer', icon="⚠️")

    if __name__ == '__main__':
        main()