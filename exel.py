import openpyxl

# Load the workbook
wb = openpyxl.Workbook()

# Select the active worksheet
ws = wb.active

# Write data to cells in the worksheet
ws['A1'] = 'Value 1'
ws['B1'] = 'Value 2'
ws['C1'] = 'Value 3'

# Save the workbook
wb.save('example.xlsx')